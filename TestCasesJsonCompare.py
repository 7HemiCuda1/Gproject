import json
import re
import openpyxl

class Block:
    # Attributes are to be immutable once set to something
    def __init__(self):
        self._name = None
        self._num = None
        self._message_lst = []
        self._outgoing_transition = None

    def set_name(self, name):
        if self._name is None:
            self._name = name
    
    def set_num(self, num):
        if self._num is None:
            self._num = num

    def set_outgoing_transition(self, transition):
        if self._outgoing_transition is None:
            self._outgoing_transition = transition

    def add_message(self, msg_id, msg_content):
        self._message_lst += [Message(msg_id, msg_content)]

    def get_name(self):
        return self._name

    def get_num(self):
        return self._num

    def get_message_lst(self):
        return self._message_lst

    def get_outgoing_transition(self):
        return self._outgoing_transition

    def __str__(self):
        title_str = "Block:\n\tID: {block_id}\n\tName: {block_name}\n\tOutgoing Transition: {outgoing}\n\t".format(\
            block_id = self._num, block_name = self._name, outgoing = self._outgoing_transition)
        message_str = "\n\t".join([str(msg) for msg in self._message_lst])
        if len(self._message_lst) == 0:
            message_str = "No Messages"
        return title_str + message_str

class Message:
    def __init__(self, msg_id, content):
        self.message_id = msg_id
        self.content = content
    def __str__(self):
        ret_str = "Message: ID = {msg_id}, content = {content}".format(msg_id = self.message_id, content = self.content)
        return ret_str
    

# May not need these unicode to byte converter methods at the end - this is to convert things to strings
# from unicode for convenience with Python

def json_load_byteified(file_handle):
    return _byteify(
        json.load(file_handle, object_hook=_byteify),
        ignore_dicts = True
    )

def json_loads_byteified(json_text):
    return _byteify(
        json.loads(json_text, object_hook=_byteify),
        ignore_dicts = True
    )

def _byteify(data, ignore_dicts = False):
    # if this is a unicode string, return its string representation
    if isinstance(data, unicode):
        return data.encode('utf-8')
    # if this is a list of values, return list of byteified values
    if isinstance(data, list):
        return [ _byteify(item, ignore_dicts=True) for item in data ]
    # if this is a dictionary, return dictionary of byteified keys and values
    # but only if we haven't already byteified it
    if isinstance(data, dict) and not ignore_dicts:
        return {
            _byteify(key, ignore_dicts=True): _byteify(value, ignore_dicts=True)
            for key, value in data.iteritems()
        }
    # if it's anything else, return it in its original form
    return data

def load_str_json_file(file_name, file_folder):
    data_file = open(file_folder + file_name)
    data = json_load_byteified(data_file)
    return data

def get_section(data, test_num, section):
    # test_num starts from 0
    return data[test_num][section]

def remove_clones(lst):
    """(list) -> list
    Returns a new list with no clones succeeding any element
    >>> remove_clones([1,1,1])
    [1]
    >>> remove_clones([2,3,4,4,3])
    [2,3,4,3]
    """
    ret_lst = [lst[0]]
    for i in range(1, len(lst)):
        prev = lst[i - 1]
        cur = lst[i]
        if prev != cur:
            ret_lst += [cur]
    return ret_lst

def merge_subsequent_messages(step_str):

    msg_pattern = "message:~(?P<msg_id>(\w+)),(?P<msg_content>(.+))"
    msg_re = re.compile(msg_pattern)
    next_is_msg = None

    step_lst = step_str.split("\n")
    str_to_add = ""
    ret_lst = []
    
    for i in range(len(step_lst) - 1):
        next_is_msg = msg_re.match(step_lst[i + 1])
        str_to_add += step_lst[i]
        if not next_is_msg:
            ret_lst += [str_to_add]
            str_to_add = ""
        else:
            str_to_add += ","
        
    ret_str = "\n".join(ret_lst)
    return ret_str

def parse_steps(data, test_num):
    raw_steps = get_section(data, test_num, "Steps")
    step_lst = [(step["action"] + ";" + step["value"]) for step in raw_steps]
    junk_translations = [("–", "-")]
    for i in range(len(step_lst)):
        for junk_translation in junk_translations:
            step_lst[i] = step_lst[i].replace(*junk_translation)
            
    unique_step_lst = remove_clones(step_lst)
    unique_step_str = "\n".join(unique_step_lst)
    
    #KEEP THE TRANSLATIONS IN ORDER - THEIR ORDER EFFECTS THE FINAL RESULT - THINK HARD BEFORE CHANGING
    # NEED TO ATTACH TRANSITIONS ONTO THE END OF THE PLAY MSG BLOCKS
    translations = [("GotoSubflow;(?P<block>(.+))\n", "Enter;\g<block>\n"), \
                    ("Enter;(?P<block_num>(\d+)) (?P<block_name>(.+))\n", "Enter;\g<block_num> \g<block_name>\n"), \
                    ("Enter;(?P<bool>(True|False))\n", "Transition;\g<bool>\n"),\
                    ("Enter;\n", "Transition;\n"), \
                    ("Press;No (?P<input>(.+))\nError;Max No(?P=input)\nEnter;Max No(?P=input)\n", \
                         "Transition;No \g<input>\n"), \
                    ("Press;(?P<number>(\d))\nEnter;(?P=number)( )?-( )?((.+))\n", \
                         "Transition;\g<number>\n"),\
                    ("SubflowReturn;ReturnFromSubflow\n", ""), \
                    ("Enter;(\D+)\n", "")]
    for translation in translations:
        old_regex = re.compile(translation[0])
        new_regex = translation[1]
        unique_step_str = re.sub(old_regex, new_regex, unique_step_str)
        
    unique_step_str = "\n".join(remove_clones(unique_step_str.split("\n")))
    
    return unique_step_str

def get_block_lst(step_str):
    step_lst = step_str.split("\n")
    
    block_pattern = "Enter;(?P<block_num>(\d+)) (?P<block_name>(.+))"
    transition_pattern = "Transition;(?P<condition>(.*))"
    message_pattern = "Play;(?P<msg_id>(\w+)) (?P<msg_content>(.+))"

    block_regex = re.compile(block_pattern)
    transition_regex = re.compile(transition_pattern)
    message_regex = re.compile(message_pattern)
    
    block_lst = []
    new_block = Block()
    for step in step_lst:
        block_match = block_regex.match(step)
        transition_match = transition_regex.match(step)
        message_match = message_regex.match(step)
        if block_match:
            block_name = block_match.group("block_name")
            block_num = block_match.group("block_num")
            if block_name and block_num == "":
                continue
            else:
                new_block.set_name(block_name)
                new_block.set_num(block_num)
        elif message_match:
            new_block.add_message(message_match.group("msg_id"), message_match.group("msg_content"))
        elif transition_match:
            outgoing_transition = transition_match.group("condition")
            if outgoing_transition:
                new_block.set_outgoing_transition(outgoing_transition)
            else:
                new_block.set_outgoing_transition("Verify")
            block_lst += [new_block]
            new_block = Block()
    return block_lst

def extract_blockpath(step_str):
    block_list = get_block_lst(step_str)
    block_id_lst = []
    for block in block_list:
        if block.get_num():
            block_id_lst += [block.get_num()]
    return "->".join(block_id_lst)

def get_relevant_blocks(block_lst):
    ret_lst = []
    for block in block_lst:
        if not (block.get_message_lst() == [] and block.get_outgoing_transition() in ["Verify", "True", "False"]):
            ret_lst += [block]
    return ret_lst

def get_variable_lst(data, test_num, tuple_headers):
    ret_lst = []
    target = get_section(data, test_num, "SymbolTable")["symbolMap"]
    for variable in target.keys():
        if any(x == variable for x in tuple_headers):
            ret_lst += [variable + " = " + get_tuple_sole_value(target[variable]["value"])]
        else:
            ret_lst += [variable + " = " + target[variable]["value"]]
    ret_lst.sort()
    return ret_lst

def get_variable_string(variable_list):
    return "\n".join(variable_list)

def get_tuple_sole_value(tup_str):
    ret_lst = tup_str.split('" ')
    for i in range(len(ret_lst)):
        ret_lst[i] = ret_lst[i].replace('"',"")
    target = ""
    for item in ret_lst:
        if "<>" not in item:
            target = item
    parsed = target.split(" = ")
    return parsed[-1].replace(" ", "")

def create_excel_document(folder_path, file_name, json_data, headers, phase, customer):
    full_path = folder_path + file_name + ".xlsx"
    wb = openpyxl.Workbook()

    # Create the summary worksheet
    create_summary_sheet(wb, json_data, headers)
    
    # Create a new worksheet for each test case
    for test_num in range(len(json_data)):
        new_ws = wb.create_sheet()
        new_ws.title = "Test Case " + str(test_num + 1)

        # Seed the worksheet with titles
        col_titles = [("A", "TC Ref"), ("B","TC Name"), ("C","Description"), \
                      ("D", "Test Steps"), ("E", "Caller Input"), \
                      ("F","Expected Result"), ("G", "Prompt Verbiage"),\
                      ("H", "Passed"), ("I", "Comments"), ("J", "POD Customer"), \
                      ("K", "Label")]
        for title in col_titles:
            new_ws[title[0] + "1"] = title[1]

        # Write the test procedure to the worksheet
        steps = parse_steps(json_data, test_num)
        block_lst = get_block_lst(steps)
        rel_block_lst = get_relevant_blocks(block_lst)
        var_lst = get_variable_lst(json_data, test_num, headers)
        var_str = get_variable_string(var_lst)

        i = 0
        for block in rel_block_lst:
            message_lst = block.get_message_lst()
            for message in message_lst:
                # Write repeated values to columns ABCDJK
                new_ws["A" + str(i + 2)] = str(test_num + 1)
                new_ws["B" + str(i + 2)] = file_name + " - TC" + str(test_num + 1)
                new_ws["C" + str(i + 2)] = var_str
                new_ws["D" + str(i + 2)] = str(i + 1)
                new_ws["J" + str(i + 2)] = customer
                new_ws["K" + str(i + 2)] = phase
            
                # Write test case procedure to columns EFG
                new_ws["E" + str(i + 2)] = "Verify"
                new_ws["F" + str(i + 2)] = str(message.message_id)
                new_ws["G" + str(i + 2)] = str(message.content)
                i += 1
            # Write the transition to the last caller input cell
            new_ws["E" + str(i + 1)] = block.get_outgoing_transition()
        
    # Save the workbook
    wb.save(full_path)

def create_summary_sheet(workbook, json_data, headers):
    # CREATING THE DIFFERENT VARIABLE LIST COULD BE MORE EFFICIENT
    # WANT TO FIGURE OUT A BETTER WAY TO ID DIFFERENT VARIABLES
    # COULD MAKE SEPERATE METHOD FOR FINDING DIFFERENT VARIABLES
    
    ws = workbook.create_sheet()
    ws.title = "Summary"
    ws["A1"] = "Test Case #"
    ws["B1"] = "Blockpath"
    ws["C1"] = "Raw Variable Settings"
    ws["D1"] = "Different Variables from Prev"
    i = 2
    for test_num in range(len(json_data)):
        parsed_steps = parse_steps(json_data, test_num)
        block_list = get_block_lst(parsed_steps)
        block_path = extract_blockpath(parsed_steps)

        var_lst = get_variable_lst(json_data, test_num, headers)
        var_str = get_variable_string(var_lst)
        
        ws["A" + str(i)] = i - 1
        ws["B" + str(i)] = block_path
        ws["C" + str(i)] = var_str
        i += 1

    for i in range(1, ws.max_row - 1):
        prev_vars = get_variable_lst(json_data, i - 1, headers)
        cur_vars = get_variable_lst(json_data, i, headers)
        prev_pairs = [prev_var.split(" = ") for prev_var in prev_vars]
        cur_pairs = [cur_var.split(" = ") for cur_var in cur_vars]

        prev_dict = {}
        for pair in prev_pairs:
            prev_dict[pair[0]] = pair[-1]

        cur_dict = {}
        for pair in cur_pairs:
            cur_dict[pair[0]] = pair[-1]

        diff_var_lst = []
        if prev_dict != cur_dict:
            for key in prev_dict.keys():
                if key in cur_dict.keys() and cur_dict[key] != prev_dict[key]:
                    diff_var_lst += [key]
            for key in cur_dict.keys():
                if key not in prev_dict.keys():
                    diff_var_lst += [key]

        diff_var_str = "\n".join(sorted(diff_var_lst))
        ws["D" + str(i + 2)] = diff_var_str.strip("\n")

#TESTING BELOW THIS LINE ======================================================
#Test
json_file_folder = """/home/jd/Downloads/"""
json_file_name1 = """Phone Cover Ingram Micro Project 1.1 TestCaseData.json"""
json_file_name2 = """Phone Cover Ingram Micro Project 1.1 TestCaseData.json"""
json_data1 = load_str_json_file(json_file_name1, json_file_folder)
json_data2 = load_str_json_file(json_file_name2, json_file_folder)

headers = ["varclosedtreatment", "dnis"]
excel_file_folder = """/home/jd/Downloads/"""
excel_file_name = """Phone Cover Ingram Micro Project 1.1 TestCaseData1"""
#create_excel_document(excel_file_folder, excel_file_name, json_data, headers, "Phase1", "BCBSNC")

